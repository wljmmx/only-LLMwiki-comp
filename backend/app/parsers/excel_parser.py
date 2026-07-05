"""Excel 解析器（基于 openpyxl）

将每个工作表转为结构化表格，提取表头和数据行。
"""
from __future__ import annotations

import hashlib

from openpyxl import load_workbook

from app.parsers.base import DocumentParser, ElementType, ParsedDocument, ParsedElement


class ExcelParser:
    format = "excel"

    def parse(self, path: str, doc_id: str) -> ParsedDocument:
        with open(path, "rb") as f:
            checksum = hashlib.sha256(f.read()).hexdigest()

        wb = load_workbook(path, data_only=True, read_only=True)
        title = wb.sheetnames[0] if wb.sheetnames else None
        elements = self._parse_workbook(wb)
        wb.close()

        return ParsedDocument(
            doc_id=doc_id, source_path=path, format="excel",
            checksum=checksum, title=title, elements=elements,
        )

    def _parse_workbook(self, wb) -> list[ParsedElement]:
        elements: list[ParsedElement] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 表格标题
            elements.append(ParsedElement(
                type=ElementType.HEADING, content=sheet_name,
                metadata={"level": 2, "sheet": sheet_name},
            ))

            # 读取所有行
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                if row and any(v is not None for v in row):
                    rows.append([str(v) if v is not None else "" for v in row])

            if not rows:
                continue

            # 限制行数，避免内存爆炸
            if len(rows) > 5000:
                rows = rows[:5000]
                elements.append(ParsedElement(
                    type=ElementType.PARAGRAPH,
                    content=f"[截断] 工作表 {sheet_name} 超过 5000 行，仅解析前 5000 行",
                    metadata={"sheet": sheet_name},
                ))

            # 输出为 TSV 格式
            table_text = "\n".join("\t".join(r) for r in rows)
            elements.append(ParsedElement(
                type=ElementType.TABLE, content=table_text,
                metadata={
                    "sheet": sheet_name,
                    "rows": len(rows),
                    "columns": len(rows[0]) if rows else 0,
                    "has_header": len(rows) > 0,
                },
            ))

        return elements